import asyncio
import json
import re
import secrets
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from fastapi import FastAPI, File, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from database.db import get_connection, hash_password, init_db, verify_password
from engine.catalogos import NATURALEZA_DATO, PROPOSITO_ANALISIS, TIPOS_IA
from engine.parsers import get_column_info, parse_file
from engine.report_gen import generate_excel_report
from engine.scorer import DQScorer

# ──────────────────────────────────────────────────────────────────────
# App
# ──────────────────────────────────────────────────────────────────────

app = FastAPI(
    title="DQ4Value",
    description="API para análisis de calidad de datos.",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()


# ──────────────────────────────────────────────────────────────────────
# In-memory stores
# ──────────────────────────────────────────────────────────────────────

_file_store: dict = {}
_analysis_store: dict = {}
_progress_store: dict = {}   # file_id -> {pct, message, done, error}

TEMP_DIR = Path(__file__).resolve().parent.parent / "temp_files"
TEMP_DIR.mkdir(exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
REPORTS_DIR  = PROJECT_ROOT / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".txt"}

# ──────────────────────────────────────────────────────────────────────
# Pydantic schemas
# ──────────────────────────────────────────────────────────────────────


class AnalyzeRequest(BaseModel):
    file_id: str
    id_column: str
    columns_config: dict[str, dict[str, dict]]
    descripcion: Optional[str] = None
    naturaleza_dato: Optional[str] = None
    proposito_analisis: Optional[str] = None
    tipo_ia: Optional[str] = None
    etiqueta: Optional[str] = None  # kept for backwards compatibility, no longer written


class LoginRequest(BaseModel):
    email: str
    password: str


class CreateUserRequest(BaseModel):
    email: str
    password: str
    nombre: str
    rol: str = "usuario"
    max_registros: int = 10000
    fecha_vencimiento: Optional[str] = None


class UpdateUserRequest(BaseModel):
    nombre: Optional[str] = None
    rol: Optional[str] = None
    max_registros: Optional[int] = None
    fecha_vencimiento: Optional[str] = None
    activo: Optional[bool] = None
    password: Optional[str] = None


# ──────────────────────────────────────────────────────────────────────
# Auth helpers
# ──────────────────────────────────────────────────────────────────────


def get_current_user(authorization: str = Header(None)) -> dict:
    if not authorization:
        raise HTTPException(status_code=401, detail="Token de autenticación requerido.")
    token = authorization.removeprefix("Bearer ").strip()
    conn = get_connection()
    now = datetime.utcnow().isoformat()
    row = conn.execute(
        """SELECT u.* FROM sesiones s
           JOIN usuarios u ON s.usuario_id = u.id
           WHERE s.token = ? AND s.fecha_expiracion > ?""",
        (token, now),
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=401, detail="Token inválido o sesión expirada.")
    return dict(row)


def require_admin(authorization: str = Header(None)) -> dict:
    user = get_current_user(authorization)
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Se requiere rol de administrador.")
    return user


# ──────────────────────────────────────────────────────────────────────
# Health
# ──────────────────────────────────────────────────────────────────────


@app.get("/health")
def health():
    return {"status": "ok", "version": "1.0.0"}


# ──────────────────────────────────────────────────────────────────────
# Auth endpoints
# ──────────────────────────────────────────────────────────────────────


@app.post("/auth/login")
def login(request: LoginRequest):
    conn = get_connection()
    user = conn.execute(
        "SELECT * FROM usuarios WHERE email = ?", (request.email,)
    ).fetchone()

    if not user or not verify_password(request.password, user["password_hash"]):
        conn.close()
        raise HTTPException(status_code=401, detail="Credenciales incorrectas.")

    if not user["activo"]:
        conn.close()
        raise HTTPException(status_code=403, detail="Usuario desactivado.")

    if user["fecha_vencimiento"]:
        if datetime.utcnow().date() > datetime.strptime(user["fecha_vencimiento"], "%Y-%m-%d").date():
            conn.close()
            raise HTTPException(status_code=403, detail="Usuario vencido.")

    # Clean only expired sessions to allow multiple concurrent sessions
    conn.execute(
        "DELETE FROM sesiones WHERE usuario_id = ? AND fecha_expiracion <= ?",
        (user["id"], datetime.utcnow().isoformat()),
    )

    token = secrets.token_urlsafe(32)
    expiry = (datetime.utcnow() + timedelta(hours=8)).isoformat()
    conn.execute(
        "INSERT INTO sesiones (usuario_id, token, fecha_expiracion) VALUES (?, ?, ?)",
        (user["id"], token, expiry),
    )
    conn.commit()
    conn.close()

    return {
        "token": token,
        "nombre": user["nombre"],
        "email": user["email"],
        "rol": user["rol"],
        "max_registros": user["max_registros"],
    }


@app.post("/auth/logout")
def logout(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Token requerido.")
    token = authorization.removeprefix("Bearer ").strip()
    conn = get_connection()
    conn.execute("DELETE FROM sesiones WHERE token = ?", (token,))
    conn.commit()
    conn.close()
    return {"message": "Sesión cerrada correctamente."}


@app.get("/auth/me")
def me(authorization: str = Header(None)):
    user = get_current_user(authorization)
    return {
        "id": user["id"],
        "nombre": user["nombre"],
        "email": user["email"],
        "rol": user["rol"],
        "max_registros": user["max_registros"],
        "fecha_vencimiento": user["fecha_vencimiento"],
    }


# ──────────────────────────────────────────────────────────────────────
# Catalogues (public)
# ──────────────────────────────────────────────────────────────────────

@app.get("/catalogos")
def get_catalogos():
    return {
        "naturaleza_dato":   NATURALEZA_DATO,
        "proposito_analisis": PROPOSITO_ANALISIS,
        "tipos_ia":          TIPOS_IA,
    }


# ──────────────────────────────────────────────────────────────────────
# File endpoints (protected)
# ──────────────────────────────────────────────────────────────────────


@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    authorization: str = Header(None),
):
    current_user = get_current_user(authorization)

    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Formato '{suffix}' no compatible. Usa: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    file_id = secrets.token_urlsafe(16)
    dest_path = TEMP_DIR / f"{file_id}{suffix}"

    try:
        content = await file.read()
        dest_path.write_bytes(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar el archivo: {e}")

    try:
        df, columns = parse_file(str(dest_path))
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ValueError as e:
        dest_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        dest_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Error interno al parsear el archivo: {e}")

    # Enforce row limit for the user
    max_reg = current_user["max_registros"]
    if len(df) > max_reg:
        dest_path.unlink(missing_ok=True)
        raise HTTPException(
            status_code=403,
            detail=f"El archivo tiene {len(df)} registros pero tu plan permite máximo {max_reg} registros.",
        )

    col_info = get_column_info(df)

    _file_store[file_id] = {
        "path": str(dest_path),
        "df": df,
        "columns": columns,
        "original_name": file.filename,
        "usuario_id": current_user["id"],
    }

    return {
        "file_id": file_id,
        "nombre_archivo": file.filename,
        "total_registros": len(df),
        "total_columnas": len(columns),
        "columnas": col_info,
    }


@app.get("/analyze/status/{file_id}")
def analyze_status(file_id: str, authorization: str = Header(None)):
    get_current_user(authorization)
    return _progress_store.get(
        file_id,
        {"pct": 0, "message": "No iniciado", "done": False, "error": None},
    )


@app.post("/analyze")
async def analyze(request: AnalyzeRequest, authorization: str = Header(None)):
    current_user = get_current_user(authorization)

    if request.file_id not in _file_store:
        raise HTTPException(
            status_code=404,
            detail=f"No se encontró el archivo con id '{request.file_id}'.",
        )
    if not request.columns_config:
        raise HTTPException(status_code=400, detail="El campo 'columns_config' no puede estar vacío.")

    stored = _file_store[request.file_id]
    df = stored["df"]
    file_path = Path(stored.get("path", ""))   # grab path now — needed for cleanup

    if request.id_column not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"La columna ID '{request.id_column}' no existe. Disponibles: {stored['columns']}",
        )
    missing = [c for c in request.columns_config if c not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas no encontradas: {missing}. Disponibles: {stored['columns']}",
        )

    # Initialise progress
    col_names  = list(request.columns_config.keys())
    total_dims = sum(len(v) for v in request.columns_config.values())
    fid        = request.file_id
    _progress_store[fid] = {"pct": 0, "message": "Iniciando análisis...", "done": False, "error": None}

    def _progress_cb(col: str, dim_name: str, done: int, total: int) -> None:
        pct     = min(5 + int((done / max(total, 1)) * 80), 84)
        col_idx = col_names.index(col) + 1 if col in col_names else "?"
        _progress_store[fid] = {
            "pct":     pct,
            "message": f"Procesando columna {col_idx} de {len(col_names)} ({col})",
            "done":    False,
            "error":   None,
        }

    def _run_sync() -> dict:
        scorer = DQScorer(df, id_col=request.id_column, progress_callback=_progress_cb)
        for col_name, dim_config in request.columns_config.items():
            scorer.configure(col_name, dim_config)
        return scorer.run_analysis()

    loop = asyncio.get_event_loop()
    try:
        try:
            results = await loop.run_in_executor(None, _run_sync)
        except Exception as e:
            print("ERROR DETALLADO:")
            traceback.print_exc()
            _progress_store[fid] = {"pct": 0, "message": str(e), "done": True, "error": str(e)}
            if isinstance(e, (ValueError, RuntimeError)):
                raise HTTPException(status_code=400, detail=str(e))
            raise HTTPException(status_code=500, detail=f"Error interno durante el análisis: {e}")

        _progress_store[fid] = {"pct": 90, "message": "Generando reporte...", "done": False, "error": None}

        # Compute summary
        summary = DQScorer.compute_summary(results)
        _analysis_store[request.file_id] = results

        # Generate persisted report
        ruta_reporte  = ""
        ruta_dashboard = ""
        _rdir = _ts = _dname = None
        try:
            _ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
            _uname = re.sub(r"[^\w\-]", "_", (current_user.get("nombre") or current_user["email"].split("@")[0]))
            _dname = re.sub(r"[^\w\-]", "_", Path(stored.get("original_name", "report")).stem)
            _rdir  = REPORTS_DIR / _uname / datetime.now().strftime("%Y-%m")
            _rdir.mkdir(parents=True, exist_ok=True)
            rfile  = _rdir / f"{_ts}_{_dname}.xlsx"
            generate_excel_report(results, str(rfile))
            ruta_reporte = str(rfile.relative_to(PROJECT_ROOT))
        except Exception as re_err:
            print(f"[report] Warning: {re_err}")

        # Generate HTML dashboard
        try:
            if _rdir and _ts and _dname:
                from engine.dashboard_gen import generate_dashboard_html
                dfile     = _rdir / f"{_ts}_{_dname}_dashboard.html"
                dash_html = generate_dashboard_html(
                    analysis_results=results,
                    filename=stored.get("original_name", ""),
                    fecha=datetime.now().strftime("%d/%m/%Y %H:%M"),
                    etiqueta=request.etiqueta or "",
                    descripcion=request.descripcion or "",
                )
                dfile.write_text(dash_html, encoding="utf-8")
                ruta_dashboard = str(dfile.relative_to(PROJECT_ROOT))
        except Exception as de_err:
            print(f"[dashboard] Warning: {de_err}")

        # Collect unique dimension names
        dims = sorted({dim for col_dims in request.columns_config.values() for dim in col_dims})

        # Persist full analysis record
        try:
            conn = get_connection()
            conn.execute(
                """INSERT INTO analisis
                   (usuario_id, file_id, nombre_archivo, total_registros, score_general,
                    descripcion, total_columnas, total_problemas,
                    dimensiones_aplicadas, ruta_reporte, ruta_dashboard, estado,
                    version_motor, naturaleza_dato, proposito_analisis, tipo_ia)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    current_user["id"],
                    request.file_id,
                    stored.get("original_name", ""),
                    results["total_registros"],
                    results["score_general"],
                    request.descripcion,
                    len(stored["columns"]),
                    results["total_problemas"],
                    json.dumps(dims),
                    ruta_reporte,
                    ruta_dashboard,
                    "completado",
                    "v2",
                    request.naturaleza_dato,
                    request.proposito_analisis,
                    request.tipo_ia,
                ),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass  # Non-critical

        _progress_store[fid] = {"pct": 100, "message": "¡Análisis completado!", "done": True, "error": None}

        return {
            "file_id": request.file_id,
            "score_general": results["score_general"],
            "total_registros": results["total_registros"],
            "total_problemas": results["total_problemas"],
            "pct_limpios": summary["pct_limpios"],
            "peor_dimension": summary["peor_dimension"],
            "scores_por_columna": results["scores_por_columna"],
        }
    finally:
        # Always delete the temp file from disk and free RAM,
        # regardless of success or error.
        try:
            if file_path and file_path.exists():
                file_path.unlink()
                print(f"[cleanup] Temp file deleted: {file_path.name}")
        except Exception as cleanup_err:
            print(f"[cleanup] Warning: could not delete temp file {file_path}: {cleanup_err}")
        _file_store.pop(request.file_id, None)


@app.get("/report/{file_id}")
def get_report(file_id: str, authorization: str = Header(None)):
    get_current_user(authorization)
    conn = get_connection()
    row  = conn.execute(
        "SELECT ruta_reporte, nombre_archivo FROM analisis WHERE file_id = ? ORDER BY id DESC LIMIT 1",
        (file_id,),
    ).fetchone()
    conn.close()
    if not row or not row["ruta_reporte"]:
        raise HTTPException(status_code=404, detail="Reporte no encontrado.")
    rfile = PROJECT_ROOT / row["ruta_reporte"]
    if not rfile.exists():
        raise HTTPException(status_code=404, detail="El archivo del reporte ya no está disponible.")
    stem = Path(row["nombre_archivo"]).stem if row["nombre_archivo"] else "reporte"
    return FileResponse(
        path=str(rfile),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"reporte_calidad_{stem}.xlsx",
    )


# ──────────────────────────────────────────────────────────────────────
# Profile endpoints
# ──────────────────────────────────────────────────────────────────────

@app.get("/profile/{file_id}")
def get_profile(file_id: str, authorization: str = Header(None)):
    from engine.profiler import profile_dataset
    get_current_user(authorization)
    if file_id not in _file_store:
        raise HTTPException(status_code=404, detail=f"Archivo '{file_id}' no encontrado.")
    df = _file_store[file_id]["df"]
    return profile_dataset(df)


class DrillThroughRequest(BaseModel):
    columna: str
    mascara: str


@app.post("/profile/{file_id}/drillthrough")
def drillthrough(file_id: str, request: DrillThroughRequest, authorization: str = Header(None)):
    from engine.profiler import drill_through
    get_current_user(authorization)
    if file_id not in _file_store:
        raise HTTPException(status_code=404, detail=f"Archivo '{file_id}' no encontrado.")
    df = _file_store[file_id]["df"]
    if request.columna not in df.columns:
        raise HTTPException(status_code=400, detail=f"Columna '{request.columna}' no existe.")
    try:
        resultado = drill_through(df, request.columna, request.mascara)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en drill-through: {e}")

    # Serializar — convertir NaN a None y tipos numpy a Python nativos
    records = []
    for row in resultado.to_dict("records"):
        clean = {}
        for k, v in row.items():
            import math
            if isinstance(v, float) and math.isnan(v):
                clean[k] = None
            elif hasattr(v, "item"):       # numpy scalar
                clean[k] = v.item()
            else:
                clean[k] = v
        records.append(clean)

    return {
        "columna": request.columna,
        "mascara": request.mascara,
        "total": len(records),
        "columnas": list(df.columns),
        "registros": records,
    }


@app.get("/profile/{file_id}/export")
def export_profile(file_id: str, authorization: str = Header(None)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from engine.profiler import profile_dataset
    get_current_user(authorization)
    if file_id not in _file_store:
        raise HTTPException(status_code=404, detail=f"Archivo '{file_id}' no encontrado.")

    df = _file_store[file_id]["df"]
    perfil = profile_dataset(df)
    resumen = perfil["resumen"]
    columnas = perfil["columnas"]

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1560A8")

    ws1.append(["Métrica", "Valor"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws1.append(["Total filas", resumen["total_filas"]])
    ws1.append(["Total columnas", resumen["total_columnas"]])
    ws1.append(["Completitud global (%)", resumen["completitud_global"]])
    ws1.append(["Filas duplicadas exactas", resumen["filas_duplicadas_exactas"]])
    ws1.append(["Tamaño en memoria (MB)", resumen["tamano_memoria_mb"]])
    ws1.append([])
    ws1.append(["Alertas detectadas", ""])
    ws1["A" + str(ws1.max_row)].font = Font(bold=True)
    for alerta in resumen["alertas"]:
        ws1.append(["⚠", alerta])
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 60

    # ── Hoja 2: Perfil por columna ───────────────────────────────────────
    ws2 = wb.create_sheet("Perfil por columna")
    headers2 = [
        "Columna", "Tipo perfil", "% Nulos", "Únicos",
        "Cardinalidad", "Min", "Max", "Promedio", "Mediana",
        "Desv. Std", "Outliers", "Sesgo",
        "Long. Prom.", "Formato detectado", "Tiene may. mezcladas",
        "Es catálogo", "Fecha mín", "Fecha máx", "Rango días",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col_name, p in columnas.items():
        row = [col_name, p.get("tipo_perfil", "")]
        row.append(p.get("pct_nulos", ""))
        row.append(p.get("total_unicos", ""))
        row.append(p.get("cardinalidad", ""))
        row.append(p.get("min", ""))
        row.append(p.get("max", ""))
        row.append(p.get("promedio", ""))
        row.append(p.get("mediana", ""))
        row.append(p.get("desviacion_std", ""))
        row.append(p.get("outliers_count", ""))
        row.append(p.get("sesgo", ""))
        row.append(p.get("longitud_promedio", ""))
        row.append(p.get("formato_detectado", ""))
        row.append(p.get("tiene_mayusculas_mezcladas", ""))
        row.append(p.get("es_catalogo", ""))
        row.append(p.get("fecha_min", ""))
        row.append(p.get("fecha_max", ""))
        row.append(p.get("rango_dias", ""))
        ws2.append(row)

    for col_letter in ["A", "B", "C", "D", "E", "N"]:
        ws2.column_dimensions[col_letter].width = 22

    # ── Hoja 3: Top valores ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Top valores")
    ws3.append(["Columna", "Valor", "Frecuencia", "Porcentaje (%)"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col_name, p in columnas.items():
        top = p.get("top_10_valores") or p.get("valores", [])
        for item in top[:10]:
            ws3.append([col_name, item.get("valor", ""), item.get("frecuencia", ""), item.get("porcentaje", "")])

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 18

    export_path = TEMP_DIR / f"perfil_{file_id}.xlsx"
    wb.save(str(export_path))
    original_name = Path(_file_store[file_id]["original_name"]).stem
    return FileResponse(
        path=str(export_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"perfil_{original_name}.xlsx",
    )


class SuggestRequest(BaseModel):
    file_id: str


@app.post("/ai/suggest")
def ai_suggest(request: SuggestRequest, authorization: str = Header(None)):
    from ai.claude_analyzer import suggest_dimensions_rules
    from engine.profiler import profile_dataset

    get_current_user(authorization)
    file_id = request.file_id

    if file_id not in _file_store:
        raise HTTPException(status_code=404, detail=f"Archivo '{file_id}' no encontrado.")

    df = _file_store[file_id]["df"]
    col_info = get_column_info(df)

    # Enrich metadata with unique-value stats needed for type-based rules
    for ci in col_info:
        col = ci["nombre"]
        ci["valores_unicos"] = int(df[col].nunique(dropna=True))
        if ci["valores_unicos"] <= 15:
            ci["top_values"] = (
                df[col].dropna().value_counts().head(10).index.astype(str).tolist()
            )
        else:
            ci["top_values"] = []

    # Pass real profile so suggestions are evidence-based
    profile: dict | None = None
    try:
        perfil  = profile_dataset(df)
        profile = perfil.get("columnas", {})
    except Exception:
        pass

    suggestions = suggest_dimensions_rules(col_info, profile=profile)
    return {
        "sugerencias":    suggestions,
        "ia_disponible":  False,
        "motor":          "rules+profile" if profile else "rules",
        "total_columnas": len(suggestions),
    }


@app.get("/issues/{file_id}")
def get_issues(file_id: str, authorization: str = Header(None)):
    get_current_user(authorization)

    # Note: _file_store is cleaned up after /analyze completes (temp file deleted).
    # Issues data lives in _analysis_store — that's the only check needed here.
    if file_id not in _analysis_store:
        raise HTTPException(status_code=404, detail=f"No se encontró análisis para '{file_id}'. Ejecuta primero POST /analyze.")

    issues_df = _analysis_store[file_id]["issues_df"]
    if issues_df.empty:
        return {"file_id": file_id, "total": 0, "issues": []}

    df = issues_df.copy()
    df = df.rename(columns={df.columns[0]: "id_registro"})
    df = df.where(df.notna(), other=None)
    return {"file_id": file_id, "total": len(df), "issues": df.to_dict(orient="records")}


# ──────────────────────────────────────────────────────────────────────
# Admin endpoints
# ──────────────────────────────────────────────────────────────────────


@app.get("/admin/usuarios")
def list_users(authorization: str = Header(None)):
    require_admin(authorization)
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, email, nombre, rol, max_registros, fecha_vencimiento, activo, fecha_creacion "
        "FROM usuarios ORDER BY fecha_creacion DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/admin/usuarios")
def create_user(request: CreateUserRequest, authorization: str = Header(None)):
    require_admin(authorization)
    if request.rol not in ("admin", "usuario"):
        raise HTTPException(status_code=400, detail="Rol debe ser 'admin' o 'usuario'.")
    conn = get_connection()
    existing = conn.execute("SELECT id FROM usuarios WHERE email = ?", (request.email,)).fetchone()
    if existing:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Ya existe un usuario con email '{request.email}'.")
    conn.execute(
        "INSERT INTO usuarios (email, password_hash, nombre, rol, max_registros, fecha_vencimiento) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (
            request.email,
            hash_password(request.password),
            request.nombre,
            request.rol,
            request.max_registros,
            request.fecha_vencimiento,
        ),
    )
    conn.commit()
    user_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return {"message": "Usuario creado.", "id": user_id}


@app.put("/admin/usuarios/{user_id}")
def update_user(user_id: int, request: UpdateUserRequest, authorization: str = Header(None)):
    require_admin(authorization)
    conn = get_connection()
    user = conn.execute("SELECT id FROM usuarios WHERE id = ?", (user_id,)).fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")

    fields, values = [], []
    if request.nombre            is not None: fields.append("nombre = ?");           values.append(request.nombre)
    if request.rol               is not None: fields.append("rol = ?");              values.append(request.rol)
    if request.max_registros     is not None: fields.append("max_registros = ?");   values.append(request.max_registros)
    if request.fecha_vencimiento is not None: fields.append("fecha_vencimiento = ?"); values.append(request.fecha_vencimiento)
    if request.activo            is not None: fields.append("activo = ?");           values.append(1 if request.activo else 0)
    if request.password          is not None:
        if len(request.password) < 6:
            conn.close()
            raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres.")
        fields.append("password_hash = ?")
        values.append(hash_password(request.password))

    if fields:
        values.append(user_id)
        conn.execute(f"UPDATE usuarios SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return {"message": "Usuario actualizado."}


@app.delete("/admin/usuarios/{user_id}")
def delete_user(user_id: int, authorization: str = Header(None)):
    admin = require_admin(authorization)
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="No puedes eliminar tu propia cuenta.")
    conn = get_connection()
    user = conn.execute("SELECT id FROM usuarios WHERE id = ?", (user_id,)).fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
    conn.execute("DELETE FROM usuarios WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"message": "Usuario eliminado."}


@app.get("/admin/usuarios/{user_id}/analisis")
def user_analyses(user_id: int, authorization: str = Header(None)):
    require_admin(authorization)
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM analisis WHERE usuario_id = ? ORDER BY fecha DESC LIMIT 50",
        (user_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ──────────────────────────────────────────────────────────────────────
# Historial endpoints
# ──────────────────────────────────────────────────────────────────────

@app.get("/historial")
def get_historial(
    fecha_desde:        Optional[str] = None,
    fecha_hasta:        Optional[str] = None,
    proposito_analisis: Optional[str] = None,
    buscar:             Optional[str] = None,
    authorization: str = Header(None),
):
    user = get_current_user(authorization)
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime("%Y-%m-%d")

    sql    = "SELECT * FROM analisis WHERE usuario_id = ? AND DATE(fecha) BETWEEN ? AND ?"
    params = [user["id"], fecha_desde, fecha_hasta]
    if proposito_analisis:
        sql += " AND proposito_analisis = ?"
        params.append(proposito_analisis)
    if buscar:
        sql += " AND (nombre_archivo LIKE ? OR descripcion LIKE ?)"
        params += [f"%{buscar}%", f"%{buscar}%"]
    sql += " ORDER BY fecha DESC"

    conn = get_connection()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/historial/stats")
def historial_stats(
    fecha_desde:        Optional[str] = None,
    fecha_hasta:        Optional[str] = None,
    proposito_analisis: Optional[str] = None,
    authorization: str = Header(None),
):
    user = get_current_user(authorization)
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime("%Y-%m-%d")

    sql    = "SELECT * FROM analisis WHERE usuario_id = ? AND DATE(fecha) BETWEEN ? AND ?"
    params = [user["id"], fecha_desde, fecha_hasta]
    if proposito_analisis:
        sql += " AND proposito_analisis = ?"
        params.append(proposito_analisis)

    conn  = get_connection()
    rows  = conn.execute(sql, params).fetchall()
    conn.close()
    rows  = [dict(r) for r in rows]
    total = len(rows)
    return {
        "total_analisis":            total,
        "score_promedio":            round(sum(r["score_general"] or 0 for r in rows) / total, 1) if total else 0,
        "total_registros_evaluados": sum(r["total_registros"] or 0 for r in rows),
        "total_datasets_distintos":  len({r["nombre_archivo"] for r in rows}),
    }


@app.get("/historial/{analisis_id}/reporte")
def download_historial_reporte(analisis_id: int, authorization: str = Header(None)):
    user = get_current_user(authorization)
    conn = get_connection()
    row  = conn.execute("SELECT * FROM analisis WHERE id = ?", (analisis_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Análisis no encontrado.")
    if row["usuario_id"] != user["id"] and user.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Sin permiso.")
    if not row["ruta_reporte"]:
        raise HTTPException(status_code=404, detail="Este análisis no tiene reporte generado.")
    rfile = PROJECT_ROOT / row["ruta_reporte"]
    if not rfile.exists():
        raise HTTPException(status_code=404, detail="El reporte de este análisis ya no está disponible.")
    stem = Path(row["nombre_archivo"]).stem if row["nombre_archivo"] else "reporte"
    return FileResponse(
        path=str(rfile),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"reporte_calidad_{stem}.xlsx",
    )


@app.get("/historial/{analisis_id}/dashboard")
def download_historial_dashboard(analisis_id: int, authorization: str = Header(None)):
    user = get_current_user(authorization)
    conn = get_connection()
    row  = conn.execute("SELECT * FROM analisis WHERE id = ?", (analisis_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Análisis no encontrado.")
    if row["usuario_id"] != user["id"] and user.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Sin permiso.")
    ruta = dict(row).get("ruta_dashboard") or ""
    if not ruta:
        raise HTTPException(status_code=404, detail="El dashboard de este análisis no está disponible.")
    dfile = PROJECT_ROOT / ruta
    if not dfile.exists():
        raise HTTPException(status_code=404, detail="El dashboard de este análisis ya no está disponible.")
    return FileResponse(
        path=str(dfile),
        media_type="text/html",
    )


@app.delete("/historial/{analisis_id}")
def delete_historial_item(analisis_id: int, authorization: str = Header(None)):
    user = get_current_user(authorization)
    conn = get_connection()
    row  = conn.execute("SELECT * FROM analisis WHERE id = ?", (analisis_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Análisis no encontrado.")
    if row["usuario_id"] != user["id"] and user.get("rol") != "admin":
        conn.close()
        raise HTTPException(status_code=403, detail="Sin permiso.")
    # Delete files from disk
    for ruta_key in ("ruta_reporte", "ruta_dashboard"):
        ruta = dict(row).get(ruta_key) or ""
        if ruta:
            try:
                f = PROJECT_ROOT / ruta
                if f.exists():
                    f.unlink()
            except Exception:
                pass
    conn.execute("DELETE FROM analisis WHERE id = ?", (analisis_id,))
    conn.commit()
    conn.close()
    return {"message": "Eliminado correctamente."}


@app.get("/admin/historial")
def admin_historial(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    etiqueta:    Optional[str] = None,
    buscar:      Optional[str] = None,
    authorization: str = Header(None),
):
    require_admin(authorization)
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime("%Y-%m-%d")

    sql    = """SELECT a.*, u.nombre as usuario_nombre, u.email as usuario_email
                FROM analisis a JOIN usuarios u ON a.usuario_id = u.id
                WHERE DATE(a.fecha) BETWEEN ? AND ?"""
    params = [fecha_desde, fecha_hasta]
    if etiqueta:
        sql += " AND a.etiqueta = ?"
        params.append(etiqueta)
    if buscar:
        sql += " AND (a.nombre_archivo LIKE ? OR a.descripcion LIKE ?)"
        params += [f"%{buscar}%", f"%{buscar}%"]
    sql += " ORDER BY a.fecha DESC"

    conn = get_connection()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]
