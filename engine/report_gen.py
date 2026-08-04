import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from engine.nombres import nombre_dual, nombre_negocio

# Paleta de colores
COLOR_GREEN  = "92D050"
COLOR_YELLOW = "FFEB84"
COLOR_RED    = "FF6B6B"
COLOR_HEADER = "2F5496"
COLOR_TITLE  = "1F3864"

FILL_GREEN  = PatternFill("solid", fgColor=COLOR_GREEN)
FILL_YELLOW = PatternFill("solid", fgColor=COLOR_YELLOW)
FILL_RED    = PatternFill("solid", fgColor=COLOR_RED)
FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_TITLE  = PatternFill("solid", fgColor=COLOR_TITLE)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _formatear_valor_if(valor, dimension: str) -> str:
    """Convierte el JSON de Isolation Forest a texto legible para Excel."""
    if dimension == 'razonabilidad' and valor and str(valor).startswith('['):
        try:
            import json
            campos = json.loads(valor)
            partes = []
            for c in campos:
                estado = '(!)' if c.get('inusual') else '(ok)'
                val = c.get('valor', 'nulo')
                partes.append(f"{c['campo']}: {val} {estado}")
            return ' | '.join(partes)
        except Exception:
            return valor
    return valor


def generate_excel_report(analysis_results: dict, output_path: str) -> str:
    """
    Genera un reporte Excel con tres pestañas:
        1. Dashboard
        2. Problemas Detallados
        3. Score por Columna

    Returns:
        Ruta absoluta del archivo generado.
    """
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja en blanco por defecto

    _build_dashboard(wb, analysis_results)
    _build_issues(wb, analysis_results)
    _build_scores_by_column(wb, analysis_results)

    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────────────────────────────
# Pestaña 1 — Dashboard
# ──────────────────────────────────────────────────────────────────────

def _build_dashboard(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Dashboard")

    score_general = results["score_general"]
    total_registros = results["total_registros"]
    total_problemas = results["total_problemas"]
    scores_por_columna = results["scores_por_columna"]

    # --- Título principal ---
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "REPORTE DE CALIDAD DE DATOS"
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = FILL_TITLE
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # --- Bloque de encabezado: veredicto + KPIs interpretados ---
    from engine.nombres import nombre_dual
    PROPOSITO_LABELS = {
        'diagnostico_general': 'Diagnóstico general', 'iniciativa_ia': 'Iniciativa de IA',
        'reporteria_bi': 'Reportería y BI', 'migracion': 'Migración de sistema',
        'integracion': 'Integración entre sistemas', 'auditoria': 'Auditoría y cumplimiento',
        'depuracion_duplicados': 'Depuración de duplicados', 'campanas': 'Campañas comerciales',
    }
    TIPO_IA_LABELS = {
        'ml_supervisado': 'ML supervisado', 'deteccion_anomalias': 'Detección de anomalías',
        'series_tiempo': 'Series de tiempo', 'segmentacion': 'Segmentación',
        'agente_generativo': 'Agente generativo', 'recomendacion': 'Recomendación',
    }
    veredicto        = results.get("veredicto", "listo")
    peor_crit        = results.get("peor_dimension_critica")
    peor_crit_sc     = results.get("peor_dimension_critica_score")
    reg_apr          = results.get("registros_aprovechables", total_registros)
    pct_apr          = results.get("pct_aprovechables", 100.0)
    pesos_origen     = results.get("pesos_origen", "proposito")
    proposito_key    = results.get("proposito_analisis", "diagnostico_general") or "diagnostico_general"
    tipo_ia_key      = results.get("tipo_ia") or ""

    if pesos_origen == 'iguales':
        pond_text = 'Ponderación: todas las dimensiones con igual peso'
    elif pesos_origen == 'manual':
        pond_text = 'Ponderación: ajustada manualmente'
    else:
        prop_label = PROPOSITO_LABELS.get(proposito_key, proposito_key)
        pond_text = f'Ponderación: perfil {prop_label}'
        if proposito_key == 'iniciativa_ia' and tipo_ia_key:
            pond_text += f' · {TIPO_IA_LABELS.get(tipo_ia_key, tipo_ia_key)}'

    VEREDICTO_LABELS = {'no_listo': 'No está listo', 'con_riesgos': 'Utilizable con reservas', 'listo': 'Listo para usar'}
    VEREDICTO_FILLS = {
        'no_listo':    PatternFill("solid", fgColor="FEE2E2"),
        'con_riesgos': PatternFill("solid", fgColor="FEF3C7"),
        'listo':       PatternFill("solid", fgColor="DCFCE7"),
    }
    VEREDICTO_FONTS = {
        'no_listo':    Font(bold=True, color="991B1B"),
        'con_riesgos': Font(bold=True, color="92400E"),
        'listo':       Font(bold=True, color="166534"),
    }

    ws.merge_cells("A3:B3")
    ws["A3"].value = "Veredicto:"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("C3:F3")
    verd_cell = ws["C3"]
    verd_cell.value = VEREDICTO_LABELS.get(veredicto, veredicto)
    verd_cell.font = VEREDICTO_FONTS.get(veredicto, Font(bold=True))
    verd_cell.fill = VEREDICTO_FILLS.get(veredicto, PatternFill())
    verd_cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[3].height = 20

    _kv(ws, 4, "Score de calidad:", f"{score_general:.1f} / 100")
    ws["B4"].font = Font(bold=True, size=14)
    ws["B4"].fill = _score_fill(score_general)

    _kv(ws, 5, "Registros aprovechables:",
        f"{reg_apr:,} de {total_registros:,}  ({pct_apr:.1f}%)")
    _kv(ws, 6, "Peor dimensión crítica:",
        f"{nombre_dual(peor_crit) if peor_crit else '—'}  (score {peor_crit_sc:.1f})" if peor_crit_sc is not None else (nombre_dual(peor_crit) if peor_crit else '—'))
    _kv(ws, 7, "Ponderación:", pond_text)

    # --- Tabla de scores por columna y dimensión ---
    ws["A9"].value = "Score por Columna y Dimensión"
    ws["A9"].font = Font(bold=True, size=11)

    header_row = 10
    headers = ["Columna", "Dimensión", "Score", "Calificación"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row = header_row + 1
    for col_name, dim_scores in scores_por_columna.items():
        for dim_name, score in dim_scores.items():
            ws.cell(row=row, column=1, value=col_name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=dim_name).border = THIN_BORDER
            score_c = ws.cell(row=row, column=3, value=round(score, 1))
            score_c.fill = _score_fill(score)
            score_c.alignment = Alignment(horizontal="center")
            score_c.border = THIN_BORDER
            cal_c = ws.cell(row=row, column=4, value=_calificacion(score))
            cal_c.fill = _score_fill(score)
            cal_c.alignment = Alignment(horizontal="center")
            cal_c.border = THIN_BORDER
            row += 1

    # ── Bloque de detalle de Similitud (si aplica) ──────────────────────
    issues_df: pd.DataFrame = results.get("issues_df")
    if issues_df is not None and not issues_df.empty and 'dimension' in issues_df.columns:
        sim_rows = issues_df[issues_df['dimension'] == 'similitud']
        # Read sim metadata from metadata_dimensiones (tuple-keyed dict)
        sim_m = {}
        for key, meta in (results.get('metadata_dimensiones') or {}).items():
            if isinstance(key, tuple) and key[1] == 'similitud' and meta:
                sim_m = meta
                break
        if not sim_rows.empty and sim_m:
            detail_start = row + 2

            ws.merge_cells(f"A{detail_start}:D{detail_start}")
            title_c = ws[f"A{detail_start}"]
            title_c.value = "Registros parecidos (similitud) — detalle"
            title_c.font = Font(bold=True, size=11, color="FFFFFF")
            title_c.fill = FILL_HEADER
            title_c.alignment = Alignment(horizontal="left")
            ws.row_dimensions[detail_start].height = 20

            def _sim_kv(r, label, value):
                ws.cell(row=r, column=1, value=label).font = Font(bold=True)
                ws.cell(row=r, column=2, value=value)

            r2 = detail_start + 1
            _sim_kv(r2,     "Grupos detectados:",            int(sim_m.get('total_grupos', 0)))
            _sim_kv(r2 + 1, "Registros involucrados:",       int(sim_m.get('total_involucrados', 0)))
            _sim_kv(r2 + 2, "Registros excedentes:",         int(sim_m.get('total_excedentes', 0)))
            _sim_kv(r2 + 3, "Duplicados exactos excluidos:", int(sim_m.get('duplicados_exactos_excluidos', 0)))
            _sim_kv(r2 + 4, "Valores vacíos excluidos:",     int(sim_m.get('placeholders_excluidos', 0)))
            _sim_kv(r2 + 5, "Algoritmo:",                    str(sim_m.get('algoritmo', '')))
            _sim_kv(r2 + 6, "Umbral:",                       f"{sim_m.get('umbral', '')}%")

    _autofit(ws, [30, 25, 12, 15])


# ──────────────────────────────────────────────────────────────────────
# Pestaña 2 — Problemas Detallados
# ──────────────────────────────────────────────────────────────────────

def _build_issues(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Problemas Detallados")
    issues_df: pd.DataFrame = results["issues_df"]

    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "PROBLEMAS DETALLADOS"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = FILL_TITLE
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if issues_df.empty:
        ws["A3"].value = "No se encontraron problemas en el dataset."
        ws["A3"].font = Font(italic=True, color="00AA00")
        return

    id_col = issues_df.columns[0]
    has_sim = (
        'dimension' in issues_df.columns
        and (issues_df['dimension'] == 'similitud').any()
        and 'grupo_id' in issues_df.columns
    )

    BASE = [id_col, 'columna', 'dimension', 'descripcion', 'valor_encontrado']

    if has_sim:
        df_w = issues_df[BASE].copy()
        df_w['Grupo']          = issues_df.get('grupo_id', pd.Series(dtype=object))
        df_w['% Similitud']    = issues_df.get('similitud_pct', pd.Series(dtype=object))
        conservar = issues_df.get('es_principal_sugerido', pd.Series(dtype=object))
        df_w['Conservar']      = conservar.map(lambda v: 'Sí' if v is True or v == True else '')
        df_w['_sg'] = df_w['Grupo'].fillna('ZZZZ')
        sorted_df = (
            df_w.sort_values(['columna', 'dimension', '_sg'])
            .drop(columns=['_sg'])
            .reset_index(drop=True)
        )
    else:
        sorted_df = issues_df[BASE].sort_values('columna').reset_index(drop=True)

    # Encabezados
    headers = list(sorted_df.columns)
    for col_idx, h in enumerate(headers, start=1):
        display_h = "Dimensión (técnica)" if h == "dimension" else h
        cell = ws.cell(row=2, column=col_idx, value=display_h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Datos
    for r_idx, row_data in sorted_df.iterrows():
        for c_idx, (col_name, value) in enumerate(row_data.items(), start=1):
            if col_name == 'valor_encontrado':
                value = _formatear_valor_if(value, row_data.get('dimension', ''))
            if col_name == 'dimension':
                value = nombre_dual(str(value))
            cell = ws.cell(row=r_idx + 3, column=c_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if col_name == 'Conservar' and value == 'Sí':
                cell.fill = PatternFill("solid", fgColor="DCFCE7")
                cell.font = Font(bold=True, color="166534")
                cell.alignment = Alignment(horizontal="center")
            if col_name == 'Grupo' and value and str(value).startswith('G'):
                cell.fill = PatternFill("solid", fgColor="EFF6FF")
                cell.font = Font(bold=True, color="1D4ED8")

    widths = [18, 20, 20, 50, 25] + ([12, 14, 10] if has_sim else [])
    _autofit(ws, widths)


# ──────────────────────────────────────────────────────────────────────
# Pestaña 3 — Score por Columna
# ──────────────────────────────────────────────────────────────────────

def _build_scores_by_column(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Score por Columna")
    scores_por_columna: dict = results["scores_por_columna"]

    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "SCORE POR COLUMNA"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = FILL_TITLE
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Recopilar todas las dimensiones usadas (unión de todos los keys)
    all_dims: list[str] = []
    for dim_scores in scores_por_columna.values():
        for d in dim_scores:
            if d not in all_dims:
                all_dims.append(d)

    # Encabezados
    headers = ["Columna"] + [nombre_dual(d) for d in all_dims] + ["Score Promedio"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Filas de datos
    for row_idx, (col_name, dim_scores) in enumerate(scores_por_columna.items(), start=3):
        ws.cell(row=row_idx, column=1, value=col_name).border = THIN_BORDER

        scores_this_row = []
        for dim_idx, dim_name in enumerate(all_dims, start=2):
            score = dim_scores.get(dim_name)
            cell = ws.cell(row=row_idx, column=dim_idx)
            if score is not None:
                cell.value = round(score, 1)
                cell.fill = _score_fill(score)
                scores_this_row.append(score)
            else:
                cell.value = "N/A"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Score promedio de la columna
        avg = round(sum(scores_this_row) / len(scores_this_row), 1) if scores_this_row else None
        avg_cell = ws.cell(row=row_idx, column=len(headers))
        avg_cell.value = avg
        avg_cell.fill = _score_fill(avg) if avg is not None else PatternFill()
        avg_cell.alignment = Alignment(horizontal="center")
        avg_cell.font = Font(bold=True)
        avg_cell.border = THIN_BORDER

    col_widths = [25] + [14] * len(all_dims) + [16]
    _autofit(ws, col_widths)


# ──────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────

def _score_fill(score: float) -> PatternFill:
    if score >= 80:
        return FILL_GREEN
    if score >= 60:
        return FILL_YELLOW
    return FILL_RED


def _calificacion(score: float) -> str:
    if score >= 80:
        return "Buena"
    if score >= 60:
        return "Regular"
    return "Crítica"


def _kv(ws, row: int, label: str, value) -> None:
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)


def _autofit(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
